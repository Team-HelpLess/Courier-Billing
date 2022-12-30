from openpyxl import load_workbook
from helper.sendings.mail import Mail


class Excel:
    def __init__(self, RecordModel, credit_record: int = 1, cash_record: int = 1) -> None:
        self.RecordModel = RecordModel

        if credit_record:
            work_book_name = "./helper/excel/CreditRecord.xlsx"
            work_credit_book = load_workbook(filename=work_book_name)
            self.delete_previous(work_book=work_credit_book)
            work_sheet_main = work_credit_book.create_sheet("Main")
            work_sheet_main.title = "Main"
            work_sheet_main = work_credit_book.active
            self.write_full_credit_record(
                work_book=work_credit_book, work_sheet_main=work_sheet_main
            )
            self.save(
                work_book=work_credit_book, work_book_name=work_book_name
            )

        if cash_record:
            work_book_name = "./helper/excel/CashRecord.xlsx"
            work_cash_book = load_workbook(filename=work_book_name)
            self.delete_previous(work_book=work_cash_book)
            work_sheet_main = work_cash_book.create_sheet("Main")
            work_sheet_main.title = "Main"
            work_sheet_main = work_cash_book.active
            self.write_full_cash_record(
                work_book=work_cash_book, work_sheet_main=work_sheet_main
            )
            self.save(work_cash_book, work_book_name=work_book_name)

        Mail.montly_mail(credit_record, cash_record)

    def write_full_credit_record(self, work_book, work_sheet_main):
        work_sheet_main["A1"].value, work_sheet_main["B1"].value = (
            "Companies",
            "No .of Records",
        )

        from_company_set = set(
            str(i[0])
            for i in self.RecordModel.objects.filter(courier_type="credit").values_list("from_company")
        )
        for index, c_name in enumerate(from_company_set):
            work_sheet_main["A" + str(index + 2)].value = c_name
            c_records = self.RecordModel.objects.filter(
                from_company=c_name, courier_type="credit"
            ).order_by("booked_date", "booked_time")
            work_sheet_main["B" + str(index + 2)].value = len(c_records)

            work_sheet_temp = work_book.create_sheet(c_name)
            work_sheet_temp.title = c_name
            work_sheet_temp["A1"].value = "Courier Number"
            work_sheet_temp["B1"].value = "Courier Type"
            work_sheet_temp["C1"].value = "Courier Company"
            work_sheet_temp["D1"].value = "From Company"
            work_sheet_temp["E1"].value = "To Company"
            work_sheet_temp["F1"].value = "To Destination"
            work_sheet_temp["G1"].value = "Courier Weight"
            work_sheet_temp["H1"].value = "Courier Rate"
            work_sheet_temp["I1"].value = "Booked Date"
            work_sheet_temp["J1"].value = "Booked Time"

            for index, c_record in enumerate(c_records):
                work_sheet_temp["A" + str(index + 2)].value = c_record.courier_number
                work_sheet_temp["B" + str(index + 2)].value = c_record.courier_type
                work_sheet_temp["C" + str(index + 2)].value = c_record.courier_company
                work_sheet_temp["D" + str(index + 2)].value = c_record.from_company
                work_sheet_temp["E" + str(index + 2)].value = c_record.to_company
                work_sheet_temp["F" + str(index + 2)].value = c_record.to_destination
                work_sheet_temp["G" + str(index + 2)].value = c_record.courier_weight
                work_sheet_temp["H" + str(index + 2)].value = c_record.courier_rate
                work_sheet_temp["I" + str(index + 2)].value = c_record.booked_date
                work_sheet_temp["J" + str(index + 2)].value = c_record.booked_time

        for index, c_name in enumerate(from_company_set):
            work_sheet_main["A" + str(index + 2)].hyperlink = f"#{c_name}!A1"

    def write_full_cash_record(self, work_book, work_sheet_main):
        work_sheet_main["A1"].value, work_sheet_main["B1"].value = (
            "Booked Dates",
            "No .of Records",
        )

        booked_date_set = set(
            str(i[0])
            for i in self.RecordModel.objects.filter(courier_type="cash").values_list("booked_date")
        )

        for index, b_date in enumerate(booked_date_set):
            work_sheet_main["A" + str(index + 2)].value = b_date
            bd_records = self.RecordModel.objects.filter(
                booked_date=b_date, courier_type="cash"
            ).order_by("booked_date", "booked_time")
            work_sheet_main["B" + str(index + 2)].value = len(bd_records)

            work_sheet_temp = work_book.create_sheet(b_date)
            work_sheet_temp.title = b_date
            work_sheet_temp["A1"].value = "Courier Number"
            work_sheet_temp["B1"].value = "Courier Type"
            work_sheet_temp["C1"].value = "Courier Company"
            work_sheet_temp["D1"].value = "From Company"
            work_sheet_temp["E1"].value = "To Company"
            work_sheet_temp["F1"].value = "To Destination"
            work_sheet_temp["G1"].value = "Courier Weight"
            work_sheet_temp["H1"].value = "Courier Rate"
            work_sheet_temp["I1"].value = "Booked Date"
            work_sheet_temp["J1"].value = "Booked Time"

            for index, bd_record in enumerate(bd_records):
                work_sheet_temp["A" + str(index + 2)].value = bd_record.courier_number
                work_sheet_temp["B" + str(index + 2)].value = bd_record.courier_type
                work_sheet_temp["C" + str(index + 2)].value = bd_record.courier_company
                work_sheet_temp["D" + str(index + 2)].value = bd_record.from_company
                work_sheet_temp["E" + str(index + 2)].value = bd_record.to_company
                work_sheet_temp["F" + str(index + 2)].value = bd_record.to_destination
                work_sheet_temp["G" + str(index + 2)].value = bd_record.courier_weight
                work_sheet_temp["H" + str(index + 2)].value = bd_record.courier_rate
                work_sheet_temp["I" + str(index + 2)].value = bd_record.booked_date
                work_sheet_temp["J" + str(index + 2)].value = bd_record.booked_time

        for index, b_date in enumerate(booked_date_set):
            work_sheet_main["A" + str(index + 2)].hyperlink = f"#{b_date}!A1"

    def save(self, work_book, work_book_name):
        work_book.save(work_book_name)

    def delete_previous(self, work_book):
        sheets = work_book.get_sheet_names()
        for i in sheets:
            sheet = work_book.get_sheet_by_name(i)
            work_book.remove_sheet(sheet)
